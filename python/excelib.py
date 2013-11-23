from openpyxl import Workbook
from openpyxl.style import Border
wb = Workbook()

def set_border(ws, cell_range):
    rows = ws.range(cell_range)
    for row in rows:
        row[0].style.borders.left.border_style = Border.BORDER_THIN
        row[-1].style.borders.right.border_style = Border.BORDER_THIN
    for c in rows[0]:
        c.style.borders.top.border_style = Border.BORDER_THIN
    for c in rows[-1]:
        c.style.borders.bottom.border_style = Border.BORDER_THIN


ws = wb.get_active_sheet()
ws.title = "Calculo"

print wb.get_sheet_names()

c= ws.cell('A1')
d= ws.cell(row=9,column=9)

for i in xrange(1,9):
    for j in xrange(1,9):
        
        ws.cell(row = i,column = j).value = i*j
        set_border(ws, "B2:I9")

wb.save('excel.xlsx')
